import PyPDF2 as pdf

from pdfminer.high_level import extract_text

import pandas as pd
import pdfplumber
import openpyxl

def remove_phrases(line, phrases_to_remove):
    for phrase in phrases_to_remove:
        line = line.replace(phrase, "")
    return line

def remove_duplicates(lines):
    unique_lines = []
    seen_lines = set()

    for line in lines:
        if line not in seen_lines:
            unique_lines.append(line)
            seen_lines.add(line)

    return unique_lines

# Define a separator line
separator_line = "*" * 50  # Adjust the number of hyphens as needed

# Phrases to remove
phrases_to_remove = ["SQ Shot Quality", "PPP Points Per Possesion =Points/Possesions", "0 0 / / #DIV/0! #DIV/0! #DIV/0! #DIV/0! #DIV/0! #DIV/0!"]

pdf_file_paths = ["SB ALA 11.21.22.pdf", "SB Idaho 11.7.22.pdf", "SB MISS 11.23.22.pdf", "SB ORE 2.5.23.pdf", "SB SEL 11.13.22.pdf",
                  "SB SEM 12.1.22.pdf", "SB UVU 11.11.22.pdf", "SB WSU 3.2.22.pdf", "SB WSU 12.20.22.pdf", "SB 1.1.23 UW.pdf", "SB 1.6.23 CU.pdf", "SB 1.15.23 UA.pdf", "SB 1.20.23 STAN.pdf",
                  "SB 1.22.23 Cal.pdf", "SB 1.27.23 USC.pdf", "SB 1.29 UCLA.pdf", "SB 2.3.23 OSU.pdf", "SB 2.10.23 UW.pdf", "SB 2.12.23 WSU .pdf", "SB 2.17.23 UA.pdf",
                  "SB 2.19 ASU.pdf", "SB 2.23 CAL.pdf", "SB 2.25 Stan .pdf", "SB 11.15.22 OKL.pdf", "SB 12.10 BYU.pdf", "SB 12.14 CU.pdf", "SB 12.17 UCR.pdf",
                  "SB 12.22 SUU.pdf", "SB 12.30 WSU.pdf"]

opp_team_abbreviations = ["ALA", "ID", "MISS", "ORE", "SEL", "SEM", "UVU", "WSU", "WSU", "UW", "CU", "UA", "STAN", "CAL", "USC", "UCLA", "OSU",
                          "UW", "WSU", "UA", "ASU", "CAL", "STAN", "OKL", "BYU", "CU", "UCR", "SUU", "WSU"]

# Create a dictionary to store the processed content for each PDF
processed_pdfs = {}

for pdf_file_path, opp_team in zip(pdf_file_paths, opp_team_abbreviations):
    all_page_text = []

    with pdfplumber.open(pdf_file_path) as pdf:
        for page_num in range(len(pdf.pages)):
            page = pdf.pages[page_num]

            # Extract text from the page
            page_text = page.extract_text()

            # Append the page text to the list
            all_page_text.append(page_text)

    # Combine the extracted text from all pages into a single string
    combined_text = '\n'.join(all_page_text)

    # Define a list to store the filtered lines
    filtered_lines = []

    include_lines = False

    # Split the text into lines
    lines = combined_text.split('\n')

    # Iterate through the lines and filter based on criteria
    for line in lines:

        # Check if line meets criteria
        if line.startswith("SMART") or line.startswith(opp_team) or line.startswith("1") or line.startswith("Utah") or line.endswith("eFG%") or line.startswith("ORB%") or line.startswith("Player"):
            # Append the line to the filtered lines list
            filtered_lines.append(line)

            #Check if the line starts with "Player"
            if line.startswith("Player"):
                # Set the flag to True to include all lines after this
                include_lines = True
            else:
                include_lines = False

        # Check if we should include this line
        if include_lines:
            # Append the line to the filtered lines list
            filtered_lines.append(line)

    # Join the filtered lines back into a single string
    filtered_page_text = '\n'.join(filtered_lines)

    # Remove specified phrases from the filtered text
    filtered_page_text = remove_phrases(filtered_page_text, phrases_to_remove)

    # Define a list to store the processed sections
    processed_sections = []

    # Split the filtered text into sections based on the header
    sections = filtered_page_text.split("Player Min FG 3FG eFG% PPP TO/40 50/50R REB/40 GPR")

    # Remove empty sections
    sections = [section.strip() for section in sections if section.strip()]

    # Initialize the title for the first section
    title = ""

    # Repeat the first two lines (PDF information) with each section
    for section in sections:
        # Extract the title from the section
        title_start = section.find("SMART BOX")
        if title_start != -1:
            title_end = section.find("\n", title_start)
            if title_end != -1:
                section_title  = section[title_start:title_end]
            else:
                section_title  = section[title_start:]
        else:
            section_title = ""

        # Include the original first two lines for each section
        processed_section = f"{separator_line}\n{title.strip()}\n{separator_line}\n{lines[0]}\n{lines[1]}\n{section}"

        processed_section_lines = [line for line in processed_section.split("\n") if line.strip()]

        # Remove duplicate rows within the section
        unique_lines = []
        seen_lines = set()
        for line in processed_section_lines:
            if line not in seen_lines:
                unique_lines.append(line)
                seen_lines.add(line)

        processed_section = "\n".join(unique_lines)

        processed_sections.append(processed_section)

    # Store the processed content in the dictionary with the PDF file name as the key
    processed_pdfs[pdf_file_path] = processed_sections

# Print or process the processed PDFs as needed
for pdf_file_path, sections in processed_pdfs.items():
    for section in sections:
        print(section)
    print(f"Processed PDF: {pdf_file_path}")

#print(processed_pdfs["SB SEM 12.1.22.pdf"])

# Create a new Excel workbook
excel_file = openpyxl.Workbook()
excel_writer = excel_file.active

# Iterate through the processed PDFs and their sections
for pdf_file_path, sections in processed_pdfs.items():
    # Create a new sheet for each PDF
    sheet = excel_file.create_sheet(title=pdf_file_path)

    # Write the sections to the Excel sheet
    for i, section in enumerate(sections):
        # Split the section into lines
        lines = section.split('\n')

        # Write each line to the Excel sheet
        for row, line in enumerate(lines, start=1):
            sheet.cell(row=row, column=i + 1, value=line)

# Remove the default sheet created by openpyxl
default_sheet = excel_file['Sheet']
excel_file.remove(default_sheet)

# Save the Excel file with all sections
excel_file.save('wbb_smartbox.xlsx')

# Close the Excel file
excel_file.close()
